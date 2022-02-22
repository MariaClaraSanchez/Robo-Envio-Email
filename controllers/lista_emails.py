import openpyxl

class Planilha:
    def __init__(self) -> None:
        """ Abre o arquivo que contém o nome e email das pessoas
        """
        self.arquivo = openpyxl.load_workbook('C:/Users/maria/Documents/Robo-Envio-Email/email.xlsx')
        self.sheet = self.arquivo['Sheet1']

    def emails(self) -> dict:
        """ Utilizada para pegar os dados para quem vai ser enviada na planilha

        Returns:
            dict: Retorna um dicionário com o nome e email para quem vai ser enviado o workmap
        """
        max_linhas = self.sheet.max_row
        max_colunas = self.sheet.max_column
        dados = {}

        for l in range(2, max_linhas+1):
            for c in range(1, max_colunas+1):
                nome = self.sheet.cell(row=l, column=1).value
                email = self.sheet.cell(row=l, column=2).value
                dados[nome] = email

        return dados